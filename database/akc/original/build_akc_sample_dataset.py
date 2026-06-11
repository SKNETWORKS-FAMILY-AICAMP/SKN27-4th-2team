"""
Build the 10-breed AKC sample dataset used in this folder.

Outputs:
  - affenpinscher_sample.csv
  - affenpinscher_sample.xlsx
  - affenpinscher_sample_table.html
  - akc_sample_build.log

Data source order:
  1. Existing affenpinscher_sample.csv rows, when present, to preserve manually
     extracted local HTML fields such as colors, markings, care text, and history.
  2. Existing akc_breeds.csv rows, when present, for legacy local crawl rows.
  3. AKC public WordPress REST and page PageMap fields for missing target breeds.

This script is intended as a reproducibility/audit helper for the sample files,
not as a large-scale crawler.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from html import escape, unescape
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent
SAMPLE_CSV = BASE_DIR / "affenpinscher_sample.csv"
SAMPLE_XLSX = BASE_DIR / "affenpinscher_sample.xlsx"
SAMPLE_HTML = BASE_DIR / "affenpinscher_sample_table.html"
LEGACY_LOCAL_CSV = BASE_DIR / "akc_breeds.csv"
LOG_FILE = BASE_DIR / "akc_sample_build.log"

TARGET_BREEDS = [
    "Affenpinscher",
    "Eurasier",
    "Hovawart",
    "Slovensky Kopov",
    "Afghan Hound",
    "Airedale Terrier",
    "Akita",
    "Alaskan Klee Kai",
    "Alaskan Malamute",
    "American Bulldog",
]

FIELDNAMES = [
    "source_file",
    "breed_name",
    "breed_url",
    "height",
    "weight",
    "life_expectancy",
    "affectionate_with_family",
    "good_with_young_children",
    "good_with_other_dogs",
    "shedding_level",
    "coat_grooming_frequency",
    "drooling_level",
    "coat_type",
    "coat_length",
    "openness_to_strangers",
    "playfulness_level",
    "watchdog_protective_nature",
    "adaptability_level",
    "trainability_level",
    "energy_level",
    "barking_level",
    "mental_stimulation_needs",
    "colors",
    "markings",
    "about_the_breed",
    "health",
    "grooming",
    "exercise",
    "training",
    "nutrition",
    "history",
]

TRAIT_MAP = {
    "Affectionate With Family": "affectionate_with_family",
    "Good With Young Children": "good_with_young_children",
    "Good With Other Dogs": "good_with_other_dogs",
    "Shedding Level": "shedding_level",
    "Coat Grooming Frequency": "coat_grooming_frequency",
    "Drooling Level": "drooling_level",
    "Coat Type": "coat_type",
    "Coat Length": "coat_length",
    "Openness To Strangers": "openness_to_strangers",
    "Playfulness Level": "playfulness_level",
    "Watchdog/Protective Nature": "watchdog_protective_nature",
    "Adaptability Level": "adaptability_level",
    "Trainability Level": "trainability_level",
    "Energy Level": "energy_level",
    "Barking Level": "barking_level",
    "Mental Stimulation Needs": "mental_stimulation_needs",
}

WP_TRAIT_MAP = {
    "activity_level": "energy_level",
    "barking_level": "barking_level",
    "coat_type": "coat_type",
    "good_with_children": "good_with_young_children",
    "good_with_dogs": "good_with_other_dogs",
    "shedding": "shedding_level",
    "trainability": "trainability_level",
}

CARE_FIELD_MAP = {
    "health": "akc_org_health",
    "grooming": "akc_org_grooming",
    "exercise": "akc_org_exercise",
    "training": "akc_org_training",
    "nutrition": "akc_org_nutrition",
}

BREED_PAGE_TRAIT_ALIASES = {
    "watchdogprotective_nature": "watchdog_protective_nature",
}


def configure_logging() -> None:
    logging.basicConfig(
        filename=LOG_FILE,
        filemode="w",
        encoding="utf-8",
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )


def clean_text(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value or "")
    value = unescape(value)
    return re.sub(r"\s+", " ", value).strip()


def blank_row() -> dict[str, str]:
    return {name: "" for name in FIELDNAMES}


def request_text(url: str) -> str:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=20) as response:
        return response.read().decode("utf-8", "replace")


def slug_from_url(url: str) -> str:
    path = urlparse(url).path.strip("/")
    return path.split("/")[-1] if path else ""


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def split_traits(traits: str, row: dict[str, str]) -> None:
    for part in (traits or "").split(";"):
        if ":" not in part:
            continue
        key, value = part.split(":", 1)
        column = TRAIT_MAP.get(key.strip())
        if column:
            row[column] = value.strip()


def row_from_legacy_csv(source_row: dict[str, str]) -> dict[str, str]:
    row = blank_row()
    row.update(
        {
            "source_file": "database/akc/akc_breeds.csv",
            "breed_name": source_row.get("breed_name", ""),
            "breed_url": source_row.get("breed_url", ""),
            "height": source_row.get("height", ""),
            "weight": source_row.get("weight", ""),
            "life_expectancy": source_row.get("life_expectancy", ""),
            "about_the_breed": source_row.get("about", ""),
            "health": source_row.get("health", ""),
            "grooming": source_row.get("grooming", ""),
            "exercise": source_row.get("exercise", ""),
            "training": source_row.get("training", ""),
            "nutrition": source_row.get("nutrition", ""),
        }
    )
    split_traits(source_row.get("traits", ""), row)
    return row


def get_breedlist() -> list[dict]:
    data = json.loads(request_text("https://www.akc.org/wp-json/json/v2/breedlist.json"))
    return list(data.values())


def pagemap_value(page_html: str, name: str) -> str:
    match = re.search(
        rf'<Attribute name="{re.escape(name)}">(.*?)</Attribute>',
        page_html,
        re.S | re.I,
    )
    if not match:
        return ""
    value = clean_text(match.group(1))
    return re.sub(r"^[^:]+:\s*", "", value)


def page_meta_breed(page_html: str) -> str:
    match = re.search(r'<meta name="og:breed" content="([^"]+)"', page_html)
    return unescape(match.group(1)).strip() if match else ""


def row_from_public_akc(item: dict) -> dict[str, str]:
    row = blank_row()
    breed_id = item.get("id")
    breed_url = item.get("url", "")
    row["source_file"] = "akc_public_rest_and_pagemap"
    row["breed_name"] = item.get("breed_name_display") or item.get("breed_name", "").replace("-", " ").title()
    row["breed_url"] = breed_url

    try:
        page_html = request_text(breed_url)
        row["breed_name"] = page_meta_breed(page_html) or row["breed_name"]
        row["height"] = pagemap_value(page_html, "height")
        row["weight"] = pagemap_value(page_html, "weight")
        row["life_expectancy"] = pagemap_value(page_html, "life_expectancy")
    except Exception as exc:
        logging.warning("Failed PageMap fetch for %s: %s", row["breed_name"], exc)

    if breed_id:
        try:
            wp = json.loads(request_text(f"https://www.akc.org/wp-json/wp/v2/breed/{breed_id}"))
            row["about_the_breed"] = clean_text((wp.get("content") or {}).get("rendered", ""))
            for source, target in WP_TRAIT_MAP.items():
                value = wp.get(source)
                if value not in (None, ""):
                    row[target] = str(value)
        except Exception as exc:
            logging.warning("Failed WP REST fetch for %s: %s", row["breed_name"], exc)

    return row


def extract_breed_page_props(page_html: str) -> dict:
    match = re.search(
        r'data-js-component="breedPage"\s+data-js-props="(.*?)"',
        page_html,
        re.S,
    )
    if not match:
        return {}
    return json.loads(unescape(match.group(1)))


def format_standard_flag(value: str | None) -> str:
    return "yes" if value == "S" else "no"


def format_colors(colors_section: dict) -> str:
    rows = []
    for color in colors_section.get("colors", []) or []:
        name = clean_text(color.get("color_long", ""))
        code = clean_text(color.get("cde_color", ""))
        standard = format_standard_flag(color.get("standard_alternate"))
        if name:
            rows.append(f"{name} (code: {code}, standard: {standard})")
    return "; ".join(rows) if rows else "None listed"


def format_markings(markings_section: dict) -> str:
    rows = []
    for marking in markings_section.get("markings", []) or []:
        name = clean_text(marking.get("markings_long", ""))
        code = clean_text(marking.get("cde_markings", ""))
        standard = format_standard_flag(marking.get("standard_alternate"))
        if name:
            rows.append(f"{name} (code: {code}, standard: {standard})")
    return "; ".join(rows) if rows else "None listed"


def apply_breed_page_props(row: dict[str, str], item: dict, overwrite: bool = False) -> dict[str, str]:
    breed_url = item.get("url") or row.get("breed_url", "")
    slug = item.get("breed_name") or slug_from_url(breed_url)
    if not breed_url or not slug:
        return row

    try:
        page_html = request_text(breed_url)
        props = extract_breed_page_props(page_html)
    except Exception as exc:
        logging.warning("Failed breedPage props fetch for %s: %s", row.get("breed_name") or slug, exc)
        return row

    breed_data = ((props.get("settings") or {}).get("breed_data") or {})
    if not breed_data:
        return row

    def set_value(column: str, value: str) -> None:
        value = clean_text(value)
        if value and (overwrite or not row.get(column)):
            row[column] = value

    basics = (breed_data.get("basics") or {}).get(slug) or {}
    standards = (breed_data.get("standards") or {}).get(slug) or {}
    description = (breed_data.get("description") or {}).get(slug) or {}
    health = (breed_data.get("health") or {}).get(slug) or {}
    history = (breed_data.get("history") or {}).get(slug) or {}
    traits = ((breed_data.get("traits") or {}).get(slug) or {}).get("traits") or {}
    colors = (breed_data.get("colors") or {}).get(slug) or {}
    markings = (breed_data.get("markings") or {}).get(slug) or {}

    set_value("source_file", "akc_breedpage_props")
    set_value("breed_name", basics.get("breed_name") or item.get("breed_name_display", ""))
    set_value("breed_url", breed_url)
    set_value("height", standards.get("height_display", ""))
    set_value("weight", standards.get("weight_display", ""))
    set_value("life_expectancy", basics.get("life_expectancy", ""))
    set_value("about_the_breed", description.get("akc_org_about") or description.get("akc_org_blurb", ""))
    set_value("history", history.get("akc_org_history", ""))
    set_value("colors", format_colors(colors))
    set_value("markings", format_markings(markings))

    for output_field, source_field in CARE_FIELD_MAP.items():
        set_value(output_field, health.get(source_field, ""))

    for trait_name, trait_data in traits.items():
        output_name = BREED_PAGE_TRAIT_ALIASES.get(trait_name, trait_name)
        if output_name not in FIELDNAMES or not isinstance(trait_data, dict):
            continue
        selected = trait_data.get("selected") or []
        if selected:
            set_value(output_name, " | ".join(clean_text(str(item)) for item in selected if item))
            continue
        score = trait_data.get("score")
        if score not in (None, "", 0, "0"):
            set_value(output_name, f"{score}/5")
        elif overwrite:
            set_value(output_name, "Not specified")

    return row


def build_rows() -> list[dict[str, str]]:
    rows_by_name: dict[str, dict[str, str]] = {}
    breed_items = {item.get("breed_name_display"): item for item in get_breedlist()}

    for row in read_csv_rows(SAMPLE_CSV):
        if row.get("breed_name") in TARGET_BREEDS:
            rows_by_name[row["breed_name"]] = {name: row.get(name, "") for name in FIELDNAMES}
            logging.info("Loaded existing sample row: %s", row["breed_name"])

    for row in read_csv_rows(LEGACY_LOCAL_CSV):
        if row.get("breed_name") in TARGET_BREEDS:
            rows_by_name[row["breed_name"]] = row_from_legacy_csv(row)
            logging.info("Loaded legacy local row: %s", row["breed_name"])

    missing = [name for name in TARGET_BREEDS if name not in rows_by_name]
    if missing:
        for name in missing:
            item = breed_items.get(name)
            if not item:
                logging.warning("Missing breed in breedlist: %s", name)
                continue
            rows_by_name[name] = row_from_public_akc(item)
            logging.info("Fetched public AKC row: %s", name)

    for name in TARGET_BREEDS:
        item = breed_items.get(name)
        if not item:
            continue
        rows_by_name[name] = apply_breed_page_props(
            rows_by_name.get(name, blank_row()),
            item,
            overwrite=True,
        )
        logging.info("Applied breedPage props: %s", name)

    return [rows_by_name[name] for name in TARGET_BREEDS if name in rows_by_name]


def write_csv(rows: list[dict[str, str]]) -> None:
    with SAMPLE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)
    logging.info("Wrote CSV: %s (%d rows)", SAMPLE_CSV, len(rows))


def write_html(rows: list[dict[str, str]]) -> None:
    compact = set(FIELDNAMES[:24]) - {"colors", "markings", "about_the_breed"}
    header_cells = "".join(f"<th>{escape(name)}</th>" for name in FIELDNAMES)
    body_rows = []

    for row in rows:
        cells = []
        for name in FIELDNAMES:
            value = row.get(name, "")
            cls = "compact" if name in compact else "long"
            if name == "breed_url" and value:
                content = f'<a href="{escape(value)}">{escape(value)}</a>'
            else:
                content = escape(value)
            cells.append(f'<td class="{cls}">{content}</td>')
        body_rows.append("<tr>" + "".join(cells) + "</tr>")

    html_doc = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>AKC Sample Breed Table</title>
  <style>
    :root {{ --blue: #003594; --line: #d9e0ea; --text: #1d2430; --muted: #5f6b7a; --bg: #f6f8fb; --cell: #ffffff; }}
    body {{ margin: 0; background: var(--bg); color: var(--text); font-family: Arial, Helvetica, sans-serif; font-size: 14px; line-height: 1.5; }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 28px 18px 40px; }}
    h1 {{ margin: 0 0 6px; font-size: 24px; line-height: 1.2; color: var(--blue); }}
    .meta {{ margin: 0 0 18px; color: var(--muted); }}
    .table-wrap {{ overflow-x: auto; background: var(--cell); border: 1px solid var(--line); border-radius: 8px; box-shadow: 0 2px 10px rgba(20, 32, 48, 0.06); }}
    table {{ width: 100%; min-width: 2200px; border-collapse: collapse; }}
    th, td {{ border-right: 1px solid var(--line); border-bottom: 1px solid var(--line); padding: 10px 12px; text-align: left; vertical-align: top; }}
    th {{ position: sticky; top: 0; z-index: 1; background: #eaf1fb; color: #13294b; font-size: 12px; white-space: nowrap; }}
    td {{ max-width: 360px; background: var(--cell); }}
    tr:last-child td {{ border-bottom: 0; }}
    th:last-child, td:last-child {{ border-right: 0; }}
    .compact {{ white-space: nowrap; }}
    .long {{ min-width: 340px; }}
  </style>
</head>
<body>
  <main>
    <h1>AKC Sample Breed Table</h1>
    <p class="meta">{len(rows)}-row table preview generated from the requested CSV columns.</p>
    <div class="table-wrap"><table><thead><tr>{header_cells}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>
  </main>
</body>
</html>
"""
    SAMPLE_HTML.write_text(html_doc, encoding="utf-8")
    logging.info("Wrote HTML preview: %s", SAMPLE_HTML)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AKC Breed Samples"
    ws.append(FIELDNAMES)
    for row in rows:
        ws.append([row.get(header, "") for header in FIELDNAMES])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="13294B")
    thin = Side(style="thin", color="D9E0EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    long_columns = {
        "colors",
        "markings",
        "about_the_breed",
        "health",
        "grooming",
        "exercise",
        "training",
        "nutrition",
        "history",
    }

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for idx, header in enumerate(FIELDNAMES, 1):
        letter = get_column_letter(idx)
        if header in long_columns:
            ws.column_dimensions[letter].width = 48
        elif header == "breed_url":
            ws.column_dimensions[letter].width = 42
        elif header == "source_file":
            ws.column_dimensions[letter].width = 24
        else:
            ws.column_dimensions[letter].width = max(14, min(28, len(header) + 2))

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 95
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    end_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="AKCBreedSamples", ref=f"A1:{end_cell}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(SAMPLE_XLSX)
    logging.info("Wrote XLSX: %s", SAMPLE_XLSX)


def main() -> None:
    configure_logging()
    logging.info("Starting AKC sample dataset build")
    logging.info("Target breeds: %s", ", ".join(TARGET_BREEDS))
    rows = build_rows()
    write_csv(rows)
    write_html(rows)
    write_xlsx(rows)
    logging.info("Completed AKC sample dataset build with %d rows", len(rows))
    print(f"Wrote {len(rows)} rows")
    print(SAMPLE_CSV)
    print(SAMPLE_XLSX)
    print(SAMPLE_HTML)
    print(LOG_FILE)


if __name__ == "__main__":
    main()
